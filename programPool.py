import openpyxl

class Excel:
    def __init__(self,excelBook,excelSheet):
        self.workbook = openpyxl.Workbook(excelBook)
        self.worksheet = self.workbook[excelSheet]

    def GetTableMax(self,rowOrColumn):
        if(rowOrColumn == "row"):
            return self.worksheet.max_row
        else:
            return self.worksheet.max_column

    def GetCellValue(self,row,column):
        return self.worksheet(row,column).value

def InsertPool_excel(excelBook,excelSheet):
    programPool_excel = {}

    def ResetRowAndColumn():
        global row
        global column
        row = 1
        column = 1

    packageList = Excel(excelBook, excelSheet)
    max_column = packageList.GetTableMax("column")
    max_row = packageList.GetTableMax("row")
    ResetRowAndColumn()
    while (column < max_column + 1):
        if (packageList.getCellValue(1, column) == "PROGRAM"):
            column_program = column
        elif (packageList.getCellValue(1, column) == "EXTENSION"):
            column_extension = column
        elif (packageList.getCellValue(1, column) == "TYPE"):
            column_type = column
        elif (packageList.getCellValue(1, column) == "REQUIRED"):
            column_required = column
        elif (packageList.getCellValue(1, column) == "TAG"):
            column_tag = column
        elif (packageList.getCellValue(1, column) == "COMPILED INTO"):
            column_compiledInto = column
        elif (packageList.getCellValue(1, column) == "CLASS"):
            column_devClass = column
        elif (packageList.getCellValue(1, column) == "LOGIC PATH"):
            column_logicPath = column
        else:
            pass
        column += 1
    ResetRowAndColumn()
    row += 1
    while (row < max_row + 1):
        programName = packageList.getCellValue(row, column_program)
        exec("programPool_excel[\"{}\"][\"Extension\"] = {}".format(programName, packageList.GetCellValue(row, column_extension)))
        exec("programPool_excel[\"{}\"][\"Type\"] = {}".format(programName, packageList.GetCellValue(row, column_type)))
        exec("programPool_excel[\"{}\"][\"Required\"] = {}".format(programName, packageList.GetCellValue(row, column_required)))
        exec("programPool_excel[\"{}\"][\"Tag\"] = {}".format(programName, packageList.GetCellValue(row, column_tag)))
        exec("programPool_excel[\"{}\"][\"CompiledInto\"] = {}".format(programName, packageList.GetCellValue(row, column_compiledInto)))
        exec("programPool_excel[\"{}\"][\"DevClass\"] = {}".format(programName, packageList.GetCellValue(row, column_devClass)))
        exec("programPool_excel[\"{}\"][\"LogicPath\"] = {}".format(programName, packageList.GetCellValue(row, column_logicPath)))
        row += 1

    return programPool_excel #TODO:Return object dict instead of json-like dict to Cpp.
