import os
import openpyxl
import shutil


class Program:
    def __init__(self, name=None, extension=None, compiledInto=None, tag=None, devClass=None, logicPath=None, type=None,
                 moduleRequired=None):
        self.Name = name
        self.Extension = extension
        self.Type = type
        self.ModuleRequired = moduleRequired
        self.Tag = tag
        self.CompiledInto = compiledInto
        self.Class = devClass
        self.LogicPath = logicPath

    def Path(self):
        return self.Name + "." + self.Extension


class File:
    def __init__(self, filePath, fileName):
        self.FilePath = filePath
        self.FileName = fileName
        if (os.path.exists(filePath) == False):
            os.makedirs(filePath)

    def Write(self, content):
        file = open(self.FilePath + "\\" + self.FileName, "w")
        file.write(content)
        file.close()


class Excel:
    def __init__(self, excelPath, excelSheet):
        self.wb = openpyxl.Workbook(excelPath)
        self.ws = self.wb[excelSheet]

    def getCellMax(self, rowOrcolumn):
        if (rowOrcolumn == "row"):
            return self.ws.max_row
        else:
            return self.ws.max_column

    def getCellValue(self, row, list):
        return self.ws(row, list).value


# Task operate
programPool = {}


def Compiled_dll(program, sourcePath, compiledPath):
    if (program.Extension == "c"):
        debug = os.popen(
            "gcc " + sourcePath + "\\" + program.Path() + " -shared -o " + compiledPath + "\\" + program.Name + ".dll")
        if (os.path.exists(compiledPath + "\\" + program.Name + ".dll") == False):
            debugLog = File(compiledPath, "error-compiled-dll-" + program.Name + ".log")
            debugLog.Write(debug.read())
            return False
        else:
            return True
    elif (program.Extension == "cpp"):
        debug = os.popen(
            "g++ " + sourcePath + "\\" + program.Path() + " -shared -o " + compiledPath + "\\" + program.Name + ".dll")
        if (os.path.exists(compiledPath + "\\" + program.Name + ".dll") == False):
            debugLog = File(compiledPath, "error-compiled-dll-" + program.Name + ".log")
            debugLog.Write(debug.read())
            return False
        else:
            return True
    elif (program.Extension == "py"):
        shutil.copyfile(sourcePath + "\\" + program.Path(),
                        compiledPath + "\\" + program.Name + "\\" + program.Name + ".py")
        pydSetUp = File(compiledPath, program.Name + "\\setup.py")
        pydSetUp.Write(
            "from distutils.core import setup\nfrom Cython.Build import cythonize\nsetup(ext_modules=cythonize(" + program.Name + ".py\"))")
        debug = os.popen("python " + compiledPath + "\\" + program.Name + "\\setup.py" + " build_ext --inplace")
        if (os.path.exists(
                compiledPath + "\\" + program.Name + "\\" + program.Name + ".cp37-win_amd64" + ".pyd") == False):  # TODO:".cp37-win_amd64" must be depended on developers' devices
            debugLog = File(compiledPath, "error-compiled-pyd-" + program.Name + ".log")
            debugLog.Write(debug.read())
            os.remove(compiledPath + "\\" + program.Name)
            return False
        else:
            shutil.copyfile(compiledPath + "\\" + program.Name + "\\" + program.Name + ".cp37-win_amd64" + ".pyd",
                            compiledPath + "\\" + program.Name + ".pyd")  # TODO:".cp37-win_amd64" must be depended on developers' devices
            os.remove(compiledPath + "\\" + program.Name)
            return True
    else:
        pass


def Compiled_exe(program, sourcePath, compiledPath):
    if (program.Extension == "c"):
        debug = os.popen("gcc -o " + compiledPath + "\\" + program.Name + ".exe " + sourcePath + "\\" + program.Path())
        if (os.path.exists(compiledPath + "\\" + program.Name + ".exe") == False):
            debugLog = File(compiledPath, "error-compiled-exe-" + program.Name + ".log")
            debugLog.Write(debug.read())
            return False
        else:
            return True
    elif (program.Extension == "cpp"):
        debug = os.popen("g++ -o " + compiledPath + "\\" + program.Name + ".exe " + sourcePath + "\\" + program.Path())
        if (os.path.exists(compiledPath + "\\" + program.Name + ".dll") == False):
            debugLog = File(compiledPath, "error-compiled-exe-" + program.Name + ".log")
            debugLog.Write(debug.read())
            return False
        else:
            return True


# Process operate
def LogoPrinter():
    print("________       .___                                  __      __            .__       .___")
    print("\_____  \    __| _/__.__. ______ ______ ____ ___.__./  \    /  \___________|  |    __| _/")
    print(" /   |   \  / __ <   |  |/  ___//  ___// __ <   |  |\   \/\/   /  _ \_  __ \  |   / __ | ")
    print("/    |    \/ /_/ |\___  |\___ \ \___ \\  ___/\___  | \        (  <_> )  | \/  |__/ /_/ | ")
    print("\_______  /\____ |/ ____/____  >____  >\___  > ____|  \__/\  / \____/|__|  |____/\____ | ")
    print("        \/      \/\/         \/     \/     \/\/            \/                         \/ ")
    print("  _________________   ____  __. __________      .__.__       .___                        ")
    print(" /   _____/\______ \ |    |/ _| \______   \__ __|__|  |    __| _/___________             ")
    print(" \_____  \  |    |  \|      <    |    |  _/  |  \  |  |   / __ |/ __ \_  __ \            ")
    print(" /        \ |    `   \    |  \   |    |   \  |  /  |  |__/ /_/ \  ___/|  | \/            ")
    print("/_______  //_______  /____|__ \  |______  /____/|__|____/\____ |\___  >__|               ")
    print("        \/         \/        \/         \/                    \/    \/                   ")


def CopyrightPrinter():
    print("\nCopyright (C) 2022 ODCIME-leoweyr.")


def CMDHelper():
    def CMDHelper_MsgFormat(msg):
        stepLen = 100 - len(str(msg))
        step = 0
        while step < stepLen:
            msg = msg + " "
            step += 1
        return "  " + msg

    print("How to use it?Just - command <option> - input \"None\" to skip")
    print(CMDHelper_MsgFormat("program add packList") + "Automatically get program config from packList.xlsx.")
    print(CMDHelper_MsgFormat(
        "program add <name> <extension> <compiledInto> <tag> <class> <logicPath> <type> <moduleRequire>") + "Insert program config.")
    print(CMDHelper_MsgFormat("program remove <name>") + "Delete program config.")
    print(CMDHelper_MsgFormat("program change <name> <characteristic> <config>") + "Change program property.")
    print(CMDHelper_MsgFormat("program list") + "Show all program message.")
    print(CMDHelper_MsgFormat("exit") + "Exit OdysseyWorld SDK Builder.")


def CMDPointer():
    cmd = input("OdysseyWorld SDK Builder>")
    if (cmd == ""):
        print("Need help?Just - help or ? - Get help message")
    elif (cmd.split()[0] == "help" or cmd.split()[0] == "?"):
        CMDHelper()
    elif (cmd.split()[0] == "program"):
        if (cmd.split()[1] == "add"):
            if (cmd.split()[2] == "packList"):
                PackageList_identity()
                print("Done!Have got program message from packList.xlsx")
            else:
                exec("{} = Program(\"{}\",\"{}\",\"{}\")".format(cmd.split()[2], cmd.split()[2], cmd.split()[3],
                                                                 cmd.split()[4]))
                if (cmd.split()[5]):
                    exec("{}.Tag = \"{}\"".format(cmd.split()[2], cmd.split()[5]))
                elif (cmd.split()[6]):
                    exec("{}.Class = \"{}\"".format(cmd.split()[2], cmd.split()[6]))
                elif (cmd.split()[7]):
                    exec("{}.LogicPath = \"{}\"".format(cmd.split()[2], cmd.split()[7]))
                elif (cmd.split()[8]):
                    exec("{}.Type = \"{}\"".format(cmd.split()[2], cmd.split()[8]))
                elif (cmd.split()[9]):
                    exec("{}.ModuleRequire = \"{}\"".format(cmd.split()[2], cmd.split()[9]))
                exec("programPool[\"{}\"] = {}".format(cmd.split()[2], cmd.split()[2]))
        elif (cmd.split()[1] == "remove"):
            exec("programPool.pop(\"{}\")".format(cmd.split()[2]))
        elif (cmd.split()[1] == "change"):
            if (cmd.split()[2] in programPool):
                if (cmd.split()[3] == "name"):
                    exec("programPool[\"{}\"].Name = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "extension"):
                    exec("programPool[\"{}\"].Extension = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "type"):
                    exec("programPool[\"{}\"].Type = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "moduleRequired"):
                    exec("programPool[\"{}\"].ModuleRequired = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "tag"):
                    exec("programPool[\"{}\"].Tag = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "compiledInto"):
                    exec("programPool[\"{}\"].CompiledInto = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "class"):
                    exec("programPool[\"{}\"].Class = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
                elif (cmd.split()[3] == "logicPath"):
                    exec("programPool[\"{}\"].LogicPath = \"{}\"".format(cmd.split()[2], cmd.split()[3]))
            else:
                print(cmd.split()[2] + "don't in programPool?Just - program add - Add it")
        else:
            print("Need help?Just - help or ? - Get help message")
    elif (cmd.split()[0] == "exit"):
        os._exit(0)
    else:
        print("Need help?Just - help or ? - Get help message")


def PackageList_identity():
    global programPool

    def ResetRowAndColumn():
        global row
        global column
        row = 1
        column = 1

    packageList = Excel("./packList.xlsx", "Sheet1")
    max_column = packageList.getCellMax("column")
    max_row = packageList.getCellMax("row")
    ResetRowAndColumn()
    while (column < max_column + 1):
        if (packageList.getCellValue(1, column) == "PROGRAM"):
            column_program = column
        elif (packageList.getCellValue(1, column) == "EXTENSION"):
            column_extension = column
        elif (packageList.getCellValue(1, column) == "TYPE"):
            column_type = column
        elif (packageList.getCellValue(1, column) == "MODULE REQUIRED"):
            column_moduleRequired = column
        elif (packageList.getCellValue(1, column) == "TAG"):
            column_tag = column
        elif (packageList.getCellValue(1, column) == "COMPILED INTO"):
            column_compiledInto = column
        elif (packageList.getCellValue(1, column) == "CLASS"):
            column_class = column
        elif (packageList.getCellValue(1, column) == "LOGIC PATH"):
            column_logicPath = column
        else:
            pass
        column += 1
    ResetRowAndColumn()
    row += 1
    while (row < max_row + 1):
        programName = packageList.getCellValue(row, column_program)
        exec("{} = Program(\"{}\")".format(programName, packageList.getCellValue(row, column_program)))
        exec("{}.Extension = \"{}\"".format(programName, packageList.getCellValue(row, column_extension)))
        exec("{}.Type = \"{}\"".format(programName, packageList.getCellValue(row, column_type)))
        exec("{}.ModuleRequired = \"{}\"".format(programName, packageList.getCellValue(row, column_moduleRequired)))
        exec("{}.Tag = \"{}\"".format(programName, packageList.getCellValue(row, column_tag)))
        exec("{}.CompiledInte = \"{}\"".format(programName, packageList.getCellValue(row, column_compiledInto)))
        exec("{}.Class = \"{}\"".format(programName, packageList.getCellValue(row, column_class)))
        exec("{}.LogicPath = \"{}\"".format(programName, packageList.getCellValue(row, column_logicPath)))
        exec("programPool[\"{}\"] = {}".format(programName.programName))
        row += 1


# Interface
def UserInterface():
    LogoPrinter()
    CopyrightPrinter()
    while True:
        CMDPointer()


if __name__ == '__main__':
    UserInterface()